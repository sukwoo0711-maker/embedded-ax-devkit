from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import ZipFile
import xml.etree.ElementTree as ET


NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def q(ns: str, tag: str) -> str:
    return f"{{{NS[ns]}}}{tag}"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def digest_items(items: list[str]) -> str:
    return sha256_text("\n".join(sorted(items)))


def col_to_index(col: str) -> int:
    value = 0
    for char in col:
        value = value * 26 + ord(char.upper()) - 64
    return value


def split_cell_ref(ref: str) -> tuple[str, int, int]:
    match = re.fullmatch(r"([A-Z]+)([0-9]+)", ref)
    if not match:
        raise ValueError(f"Unsupported cell reference: {ref}")
    col, row = match.group(1), int(match.group(2))
    return col, row, col_to_index(col)


@dataclass(frozen=True)
class RangeRef:
    ref: str
    start_col: str
    start_row: int
    start_col_index: int
    end_col: str
    end_row: int
    end_col_index: int

    @property
    def width(self) -> int:
        return self.end_col_index - self.start_col_index + 1

    @property
    def height(self) -> int:
        return self.end_row - self.start_row + 1

    @property
    def anchor(self) -> str:
        return f"{self.start_col}{self.start_row}"


def parse_range_ref(ref: str) -> RangeRef:
    if ":" in ref:
        start, end = ref.split(":", 1)
    else:
        start = end = ref
    start_col, start_row, start_col_index = split_cell_ref(start)
    end_col, end_row, end_col_index = split_cell_ref(end)
    return RangeRef(ref, start_col, start_row, start_col_index, end_col, end_row, end_col_index)


def normalize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def load_shared_strings(zip_file: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zip_file.namelist():
        return []
    root = ET.fromstring(zip_file.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for item in root.findall(q("main", "si")):
        values.append("".join(node.text or "" for node in item.iter(q("main", "t"))))
    return values


def workbook_sheets(zip_file: ZipFile) -> list[dict[str, str]]:
    workbook = ET.fromstring(zip_file.read("xl/workbook.xml"))
    rel_root = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))
    rels = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rel_root.findall(q("pkgrel", "Relationship"))
    }

    sheets: list[dict[str, str]] = []
    for sheet in workbook.find(q("main", "sheets")).findall(q("main", "sheet")):
        rel_id = sheet.attrib[q("rel", "id")]
        target = rels[rel_id]
        if target.startswith("/"):
            path = target.lstrip("/")
        else:
            path = "xl/" + target.lstrip("/")
        sheets.append({"name": sheet.attrib["name"], "path": path})
    return sheets


def cell_value(cell: ET.Element, shared_strings: list[str]) -> tuple[str, bool]:
    formula = cell.find(q("main", "f"))
    is_formula = formula is not None
    if is_formula:
        return "=" + (formula.text or ""), True

    value = cell.find(q("main", "v"))
    if value is not None:
        raw = value.text or ""
        if cell.attrib.get("t") == "s":
            try:
                return shared_strings[int(raw)], False
            except (ValueError, IndexError):
                return raw, False
        return raw, False

    inline = cell.find(q("main", "is"))
    if inline is not None:
        return "".join(node.text or "" for node in inline.iter(q("main", "t"))), False

    return "", is_formula


def merge_lookup(merge_ranges: list[RangeRef]) -> dict[str, RangeRef]:
    return {item.anchor: item for item in merge_ranges}


def inspect_sheet(
    zip_file: ZipFile,
    sheet: dict[str, str],
    shared_strings: list[str],
    include_values: bool,
) -> dict[str, Any]:
    root = ET.fromstring(zip_file.read(sheet["path"]))
    dimension_el = root.find(q("main", "dimension"))
    dimension = dimension_el.attrib.get("ref", "") if dimension_el is not None else ""

    merge_el = root.find(q("main", "mergeCells"))
    merge_ranges = (
        [parse_range_ref(item.attrib["ref"]) for item in merge_el.findall(q("main", "mergeCell"))]
        if merge_el is not None
        else []
    )
    merges_by_anchor = merge_lookup(merge_ranges)

    row_nonempty: Counter[int] = Counter()
    row_merge_groups: defaultdict[int, int] = defaultdict(int)
    wide_merged_headers = []
    for item in merge_ranges:
        if item.width >= 4:
            row_merge_groups[item.start_row] += 1
            wide_merged_headers.append(
                {
                    "range": item.ref,
                    "row": item.start_row,
                    "width": item.width,
                    "height": item.height,
                }
            )

    cell_fingerprints: list[str] = []
    notable_cells: list[dict[str, Any]] = []
    formula_count = 0
    styled_cell_count = 0
    max_row = 0
    max_col = 0

    sheet_data = root.find(q("main", "sheetData"))
    if sheet_data is not None:
        for row in sheet_data.findall(q("main", "row")):
            row_number = int(row.attrib.get("r", "0"))
            max_row = max(max_row, row_number)
            for cell in row.findall(q("main", "c")):
                ref = cell.attrib["r"]
                col, cell_row, col_index = split_cell_ref(ref)
                max_row = max(max_row, cell_row)
                max_col = max(max_col, col_index)
                if cell.attrib.get("s"):
                    styled_cell_count += 1

                value, is_formula = cell_value(cell, shared_strings)
                if is_formula:
                    formula_count += 1
                if value == "" and not is_formula:
                    continue

                row_nonempty[cell_row] += 1
                cell_fingerprints.append(f"{ref}:{sha256_text(value)}")

                anchored_merge = merges_by_anchor.get(ref)
                if include_values or anchored_merge or len(notable_cells) < 20:
                    entry: dict[str, Any] = {
                        "ref": ref,
                        "row": cell_row,
                        "col": col,
                        "value_hash": sha256_text(value),
                    }
                    if include_values:
                        entry["value"] = value
                    if anchored_merge is not None:
                        entry["merged_range"] = anchored_merge.ref
                        entry["merged_width"] = anchored_merge.width
                        entry["merged_height"] = anchored_merge.height
                    notable_cells.append(entry)

    merged_ranges = [item.ref for item in merge_ranges]
    sparse_rows = [
        {"row": row, "nonempty_cells": count}
        for row, count in sorted(row_nonempty.items())
        if count >= 3
    ][:20]

    risk_flags = []
    if max_col >= 30:
        risk_flags.append("wide-sheet")
    if len(merge_ranges) >= 50:
        risk_flags.append("heavy-merged-cells")
    if len(cell_fingerprints) > 0 and max_row * max(max_col, 1) / len(cell_fingerprints) >= 10:
        risk_flags.append("sparse-lookup-grid")
    if any(count >= 2 for count in row_merge_groups.values()):
        risk_flags.append("side-by-side-column-groups")

    return {
        "name": sheet["name"],
        "dimension": dimension,
        "max_row_seen": max_row,
        "max_col_seen": max_col,
        "nonempty_cell_count": len(cell_fingerprints),
        "formula_count": formula_count,
        "styled_cell_count": styled_cell_count,
        "merged_cell_count": len(merge_ranges),
        "merge_range_digest": digest_items(merged_ranges),
        "cell_fingerprint_digest": digest_items(cell_fingerprints),
        "wide_merged_header_count": len(wide_merged_headers),
        "multi_group_header_rows": [
            {"row": row, "wide_merged_groups": count}
            for row, count in sorted(row_merge_groups.items())
            if count >= 2
        ],
        "sample_merge_ranges": merged_ranges[:20],
        "sample_sparse_rows": sparse_rows,
        "sample_wide_merged_headers": wide_merged_headers[:20],
        "notable_cells": notable_cells[:40],
        "llm_risk_flags": risk_flags,
    }


def inspect_workbook(path: Path, source_url: str | None, include_values: bool) -> dict[str, Any]:
    workbook_bytes = path.read_bytes()
    with ZipFile(path) as zip_file:
        shared_strings = load_shared_strings(zip_file)
        sheets = workbook_sheets(zip_file)
        sheet_profiles = [
            inspect_sheet(zip_file, sheet, shared_strings, include_values)
            for sheet in sheets
        ]

    return {
        "schema_version": 1,
        "tool": "scripts/inspect_excel_layout.py",
        "source": {
            "filename": path.name,
            "source_url": source_url,
            "sha256": sha256_bytes(workbook_bytes),
        },
        "workbook": {
            "sheet_count": len(sheet_profiles),
            "sheet_names": [sheet["name"] for sheet in sheet_profiles],
            "shared_string_count": len(shared_strings),
        },
        "sheets": sheet_profiles,
    }


def validate_with_openpyxl(path: Path, profile: dict[str, Any]) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False, read_only=False)
    checks: list[dict[str, Any]] = []

    checks.append(
        {
            "name": "workbook.sheet_count",
            "expected": profile["workbook"]["sheet_count"],
            "actual": len(workbook.sheetnames),
            "passed": profile["workbook"]["sheet_count"] == len(workbook.sheetnames),
        }
    )
    checks.append(
        {
            "name": "workbook.sheet_names",
            "expected": profile["workbook"]["sheet_names"],
            "actual": workbook.sheetnames,
            "passed": profile["workbook"]["sheet_names"] == workbook.sheetnames,
        }
    )

    profile_by_name = {sheet["name"]: sheet for sheet in profile["sheets"]}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        expected = profile_by_name[sheet_name]

        merged_ranges = sorted(str(item) for item in worksheet.merged_cells.ranges)
        cell_fingerprints = []
        formula_count = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                value = normalize_value(cell.value)
                if value.startswith("="):
                    formula_count += 1
                cell_fingerprints.append(f"{cell.coordinate}:{sha256_text(value)}")

        sheet_checks = [
            ("dimension", expected["dimension"], worksheet.calculate_dimension()),
            ("merged_cell_count", expected["merged_cell_count"], len(merged_ranges)),
            ("merge_range_digest", expected["merge_range_digest"], digest_items(merged_ranges)),
            ("nonempty_cell_count", expected["nonempty_cell_count"], len(cell_fingerprints)),
            ("cell_fingerprint_digest", expected["cell_fingerprint_digest"], digest_items(cell_fingerprints)),
            ("formula_count", expected["formula_count"], formula_count),
        ]
        for field, expected_value, actual_value in sheet_checks:
            checks.append(
                {
                    "name": f"sheet.{sheet_name}.{field}",
                    "expected": expected_value,
                    "actual": actual_value,
                    "passed": expected_value == actual_value,
                }
            )

    return checks


def validate_with_xml(path: Path, profile: dict[str, Any]) -> list[dict[str, Any]]:
    actual = inspect_workbook(path, profile["source"].get("source_url"), include_values=False)
    checks: list[dict[str, Any]] = []
    checks.append(
        {
            "name": "workbook.sha256",
            "expected": profile["source"]["sha256"],
            "actual": actual["source"]["sha256"],
            "passed": profile["source"]["sha256"] == actual["source"]["sha256"],
        }
    )
    checks.append(
        {
            "name": "workbook.sheet_count",
            "expected": profile["workbook"]["sheet_count"],
            "actual": actual["workbook"]["sheet_count"],
            "passed": profile["workbook"]["sheet_count"] == actual["workbook"]["sheet_count"],
        }
    )
    expected_by_name = {sheet["name"]: sheet for sheet in profile["sheets"]}
    actual_by_name = {sheet["name"]: sheet for sheet in actual["sheets"]}
    for name, expected in expected_by_name.items():
        current = actual_by_name.get(name)
        if current is None:
            checks.append({"name": f"sheet.{name}.exists", "passed": False})
            continue
        for field in [
            "dimension",
            "merged_cell_count",
            "merge_range_digest",
            "nonempty_cell_count",
            "cell_fingerprint_digest",
            "formula_count",
        ]:
            checks.append(
                {
                    "name": f"sheet.{name}.{field}",
                    "expected": expected[field],
                    "actual": current[field],
                    "passed": expected[field] == current[field],
                }
            )
    return checks


def validate_workbook(path: Path, profile_path: Path, validator: str) -> dict[str, Any]:
    profile = json.loads(profile_path.read_text(encoding="utf-8"))
    workbook_hash = sha256_bytes(path.read_bytes())
    hash_check = {
        "name": "workbook.sha256",
        "expected": profile["source"]["sha256"],
        "actual": workbook_hash,
        "passed": profile["source"]["sha256"] == workbook_hash,
    }

    active_validator = validator
    if validator == "auto":
        try:
            import openpyxl  # noqa: F401

            active_validator = "openpyxl"
        except Exception:
            active_validator = "xml"

    if active_validator == "openpyxl":
        checks = [hash_check, *validate_with_openpyxl(path, profile)]
    elif active_validator == "xml":
        checks = [*validate_with_xml(path, profile)]
    else:
        raise ValueError(f"Unsupported validator: {validator}")

    passed = all(check.get("passed", False) for check in checks)
    return {
        "schema_version": 1,
        "tool": "scripts/inspect_excel_layout.py",
        "source_filename": path.name,
        "profile_filename": profile_path.name,
        "validator": active_validator,
        "passed": passed,
        "check_count": len(checks),
        "failed_checks": [check for check in checks if not check.get("passed", False)],
        "checks": checks,
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Inspect Excel layout features that simple CSV conversion loses, then loopback-validate the profile."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect_parser = subparsers.add_parser("inspect", help="Create a layout-preserving workbook profile.")
    inspect_parser.add_argument("workbook", type=Path)
    inspect_parser.add_argument("--out", type=Path, required=True)
    inspect_parser.add_argument("--source-url", default=None)
    inspect_parser.add_argument(
        "--include-values",
        action="store_true",
        help="Include raw cell values. Leave disabled for private workbooks.",
    )

    validate_parser = subparsers.add_parser("validate", help="Loopback-check a profile against the source workbook.")
    validate_parser.add_argument("workbook", type=Path)
    validate_parser.add_argument("--profile", type=Path, required=True)
    validate_parser.add_argument("--out", type=Path, required=True)
    validate_parser.add_argument("--validator", choices=["auto", "openpyxl", "xml"], default="auto")

    args = parser.parse_args()

    if args.command == "inspect":
        profile = inspect_workbook(args.workbook, args.source_url, args.include_values)
        write_json(args.out, profile)
        sheet_summary = ", ".join(
            f"{sheet['name']}: {sheet['dimension']}, merges={sheet['merged_cell_count']}, cells={sheet['nonempty_cell_count']}"
            for sheet in profile["sheets"]
        )
        print(f"Wrote profile: {args.out}")
        print(sheet_summary)
        return 0

    if args.command == "validate":
        report = validate_workbook(args.workbook, args.profile, args.validator)
        write_json(args.out, report)
        print(f"Wrote validation report: {args.out}")
        print(f"passed={report['passed']} checks={report['check_count']} validator={report['validator']}")
        if not report["passed"]:
            for check in report["failed_checks"]:
                print(f"FAILED {check['name']}: expected={check.get('expected')} actual={check.get('actual')}")
            return 1
        return 0

    return 2


if __name__ == "__main__":
    raise SystemExit(main())
